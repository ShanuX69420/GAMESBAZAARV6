"""Build and convert the bulk-listing spreadsheet.

Runs on the local PC (needs openpyxl); the server never touches xlsx.

  python tools/listings_sheet.py build catalog.json GamesBazaar-Listings.xlsx
  python tools/listings_sheet.py convert GamesBazaar-Listings.xlsx rows.json

`catalog.json` comes from `manage.py export_catalog`; `rows.json` feeds
`manage.py import_listings`.
"""
import json
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLUMNS = [
    ('game', 28),
    ('category', 22),
    ('option', 22),
    ('title', 45),
    ('description', 50),
    ('price', 10),
    ('quantity', 10),
    ('delivery_time', 14),
    ('instant_delivery', 16),
    ('delivery_codes', 40),
    ('delivery_instructions', 50),
    ('filters', 40),
]
DELIVERY_TIMES = ['2-3 Minutes', '5 Minutes', '10-15 Minutes', '15-30 Minutes',
                  '30-60 Minutes', '1-2 Hours', '2-6 Hours', '6-12 Hours',
                  '12-24 Hours', '1-3 Days']
MAX_DATA_ROW = 5000

EXAMPLES = [
    {
        'game': 'PUBG Mobile', 'category': 'Accounts',
        'title': 'Conqueror Account Lvl 70 — 5 Mythics', 'description':
            'Conqueror rank account, level 70, 5 mythic outfits, all maps unlocked.',
        'price': '25000', 'quantity': '1', 'delivery_time': '1-2 Hours',
        'instant_delivery': 'no',
        'delivery_instructions': 'After purchase, send your email in order chat. '
                                 'Change the password as soon as you receive the account.',
        'filters': 'Rank=Conqueror; Level=61-80',
        '_note': 'Normal listing: you deliver manually within the delivery time.',
    },
    {
        'game': 'Steam', 'category': 'Gift Cards',
        'title': 'Steam Wallet Code PKR 1000', 'description':
            'Original Steam wallet code, region free. Delivered instantly.',
        'price': '1150', 'quantity': '', 'delivery_time': '',
        'instant_delivery': 'yes',
        'delivery_codes': 'AAAA-BBBB-CCCC\nDDDD-EEEE-FFFF\nGGGG-HHHH-IIII',
        'delivery_instructions': 'Redeem in Steam: Games menu > Redeem a Steam Wallet Code.',
        'filters': '',
        '_note': 'Instant delivery: one code per line; 3 lines = 3 in stock '
                 '(quantity is counted automatically). Buyer gets the next unused line.',
    },
    {
        'game': 'PUBG Mobile', 'category': 'UC Top-up', 'option': '60 UC',
        'description': '', 'price': '786', 'quantity': '10',
        'delivery_time': '1-2 Hours', 'instant_delivery': 'no',
        'delivery_instructions': 'Send your Player ID in the order chat. '
                                 'We never need your password.',
        'filters': '',
        '_note': 'Offer-style category: fill "option" with one of the preset options '
                 '(see Reference sheet), leave title empty — it is set automatically.',
    },
]


def build(catalog_path, out_path):
    with open(catalog_path, encoding='utf-8') as fh:
        catalog = json.load(fh)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2E7D32')
    wrap = Alignment(wrap_text=True, vertical='top')

    # ── Listings sheet ──
    ws = wb.active
    ws.title = 'Listings'
    for idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = 'A2'

    col = {name: i + 1 for i, (name, _) in enumerate(COLUMNS)}
    dv_time = DataValidation(
        type='list', formula1='"' + ','.join(DELIVERY_TIMES) + '"', allow_blank=True)
    dv_yesno = DataValidation(type='list', formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv_time)
    ws.add_data_validation(dv_yesno)
    time_col = get_column_letter(col['delivery_time'])
    auto_col = get_column_letter(col['instant_delivery'])
    dv_time.add(f'{time_col}2:{time_col}{MAX_DATA_ROW}')
    dv_yesno.add(f'{auto_col}2:{auto_col}{MAX_DATA_ROW}')

    # ── Reference sheet ──
    ref = wb.create_sheet('Reference')
    ref_headers = ['game', 'category', 'type', 'instant delivery allowed?',
                   'options (offer categories)', 'filters and allowed values']
    for idx, name in enumerate(ref_headers, start=1):
        cell = ref.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
    for width, letter in zip([28, 22, 12, 22, 40, 80], 'ABCDEF'):
        ref.column_dimensions[letter].width = width
    ref.freeze_panes = 'A2'

    for r, gc in enumerate(catalog, start=2):
        filters_text = '\n'.join(
            f"{f['name']}: " + ' / '.join(o['label'] for o in f['options'])
            for f in gc['filters']
        )
        ref.cell(row=r, column=1, value=gc['game'])
        ref.cell(row=r, column=2, value=gc['category'])
        ref.cell(row=r, column=3, value='offer' if gc['listing_mode'] == 'offer' else 'standard')
        ref.cell(row=r, column=4, value='yes' if gc['allow_auto_delivery'] else 'no')
        ref.cell(row=r, column=5, value=', '.join(o['name'] for o in gc['options']))
        cell = ref.cell(row=r, column=6, value=filters_text)
        cell.alignment = wrap

    # ── Examples sheet ──
    ex = wb.create_sheet('Examples')
    ex_cols = COLUMNS + [('what this example shows', 60)]
    for idx, (name, width) in enumerate(ex_cols, start=1):
        cell = ex.cell(row=1, column=idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ex.column_dimensions[get_column_letter(idx)].width = width
    for r, example in enumerate(EXAMPLES, start=2):
        for idx, (name, _) in enumerate(COLUMNS, start=1):
            cell = ex.cell(row=r, column=idx, value=example.get(name, ''))
            cell.alignment = wrap
        ex.cell(row=r, column=len(COLUMNS) + 1, value=example['_note']).alignment = wrap

    wb.save(out_path)
    print(f'Template with {len(catalog)} game/category combos -> {out_path}')


def convert(xlsx_path, out_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Listings']
    headers = [str(c.value or '').strip().casefold() for c in ws[1]]

    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = {}
        for header, cell in zip(headers, row):
            if header:
                values[header] = '' if cell.value is None else str(cell.value)
        if not any(v.strip() for v in values.values()):
            continue
        values['row'] = r
        rows.append(values)

    with open(out_path, 'w', encoding='utf-8') as fh:
        json.dump(rows, fh, indent=2, ensure_ascii=False)
    print(f'{len(rows)} filled rows -> {out_path}')


def main():
    if len(sys.argv) != 4 or sys.argv[1] not in ('build', 'convert'):
        print(__doc__)
        sys.exit(1)
    {'build': build, 'convert': convert}[sys.argv[1]](sys.argv[2], sys.argv[3])


if __name__ == '__main__':
    main()
